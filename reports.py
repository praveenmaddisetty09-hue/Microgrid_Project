"""
Enhanced Report Generation Module for Smart Microgrid Manager Pro
Generates comprehensive PDF, HTML, Excel, and text reports with professional formatting.
"""

import io
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots

# Import branding module for logo integration
from branding import (
    COMPANY_BRANDING, get_report_header, get_footer_html, 
    get_branding_css, SVG_LOGO, SVG_LOGO_SMALL
)

# Try to import Excel support
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ReportLab imports (optional - with fallback)
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def set_company_branding(branding: Dict):
    """Set company branding settings."""
    COMPANY_BRANDING.update(branding)


class ExcelReportGenerator:
    """Generate professional Excel reports with formatting."""
    
    def __init__(self, branding: Dict = None):
        self.branding = COMPANY_BRANDING.copy()
        if branding:
            self.branding.update(branding)
    
    def generate_report(self, df_results: pd.DataFrame, summary: Dict,
                      parameters: Dict) -> Tuple[bytes, str]:
        """Generate Excel report."""
        
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV
            csv = df_results.to_csv(index=False).encode('utf-8')
            return csv, 'text/csv'
        
        wb = Workbook()
        
        # Remove default sheet
        ws = wb.active
        ws.title = "Summary"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        secondary_fill = PatternFill(start_color="48BB78", end_color="48BB78", fill_type="solid")
        title_font = Font(bold=True, size=16, color="1A365D")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"{self.branding['company_name']}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:E2')
        ws['A2'] = f"{self.branding['report_title']} - {datetime.now().strftime('%Y-%m-%d')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Executive Summary Section
        ws['A4'] = "EXECUTIVE SUMMARY"
        ws['A4'].font = Font(bold=True, size=14, color="48BB78")
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Cost (₹)', round(summary.get('total_cost', 0), 2)],
            ['CO₂ Emissions (kg)', round(summary.get('total_emissions', 0), 2)],
            ['Grid Usage (kWh)', round(summary.get('total_grid_usage', 0), 2)],
            ['Renewable %', f"{summary.get('renewable_percentage', 0)}%"]
        ]
        
        for i, row in enumerate(summary_data, start=5):
            for j, value in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                if i == 5:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
        
        # Parameters Section
        ws['A11'] = "SYSTEM PARAMETERS"
        ws['A11'].font = Font(bold=True, size=14, color="48BB78")
        
        param_data = [
            ['Parameter', 'Value'],
            ['Battery Capacity', f"{parameters.get('battery_size', 100)} kWh"],
            ['Base Price', f"₹{parameters.get('base_price', 5)}/unit"],
            ['Peak Price', f"₹{parameters.get('peak_price', 15)}/unit"],
            ['Grid Safety Limit', f"{parameters.get('grid_safety_limit', 50)} kW"],
            ['Carbon Intensity', f"{parameters.get('carbon_intensity', 0.82)} kg CO₂/kWh"]
        ]
        
        for i, row_data in enumerate(param_data, start=12):
            for j, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=j, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                if i == 12:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
        
        # Hourly Data Sheet
        ws_hourly = wb.create_sheet("Hourly Data")
        
        # Add headers
        for col, header in enumerate(df_results.columns, start=1):
            cell = ws_hourly.cell(row=1, column=col, value=header if header != 'Hour' else 'Hour')
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_idx, row_data in enumerate(df_results.values, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_hourly.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # Adjust column widths
        for col in range(1, min(len(df_results.columns) + 1, 7)):
            ws_hourly.column_dimensions[chr(64 + col)].width = 18
        
        # Recommendations Sheet
        ws_rec = wb.create_sheet("Recommendations")
        ws_rec['A1'] = "💡 RECOMMENDATIONS"
        ws_rec['A1'].font = Font(bold=True, size=14, color="48BB78")
        
        recommendations = self._generate_recommendations(summary, parameters)
        for i, rec in enumerate(recommendations, start=3):
            cell = ws_rec.cell(row=i, column=1, value=f"• {rec}")
            cell.alignment = Alignment(wrap_text=True)
        
        ws_rec.column_dimensions['A'].width = 100
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"microgrid_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    def _generate_recommendations(self, summary: Dict, parameters: Dict) -> List[str]:
        """Generate recommendations."""
        recommendations = []
        renewable_pct = summary.get('renewable_percentage', 0)
        total_cost = summary.get('total_cost', 0)
        total_emissions = summary.get('total_emissions', 0)
        grid_usage = summary.get('total_grid_usage', 0)
        
        if renewable_pct < 50:
            recommendations.append("Consider increasing renewable energy capacity to reduce carbon footprint.")
        elif renewable_pct < 70:
            recommendations.append("Good renewable penetration! Explore battery storage expansion.")
        else:
            recommendations.append("Excellent renewable energy integration! Maintain current configuration.")
        
        if total_cost > 2000:
            recommendations.append("High energy costs detected. Review peak hour usage and consider load shifting.")
        
        if total_emissions > 100:
            recommendations.append("Consider carbon offset programs or additional renewable capacity.")
        
        if grid_usage > 500:
            recommendations.append("High grid dependency. Evaluate increasing battery capacity for better autonomy.")
        
        return recommendations


class EnhancedReportGenerator:
    """Enhanced report generator with multiple formats."""
    
    def __init__(self):
        self.excel_gen = ExcelReportGenerator()
    
    def generate_report(self, df_results: pd.DataFrame, summary: Dict,
                      parameters: Dict, report_format: str = 'html') -> Tuple[bytes, str]:
        """Generate a report in the specified format."""
        
        if report_format == 'excel':
            return self.excel_gen.generate_report(df_results, summary, parameters)
        elif report_format == 'text':
            return generate_text_report(df_results, summary, parameters), 'text/plain'
        else:
            return generate_html_report(df_results, summary, parameters), 'text/html'


# Legacy PDF Report Generator (kept for compatibility)
class PDFReportGenerator:
    """Generate comprehensive PDF reports for microgrid analysis."""
    
    def __init__(self):
        if not REPORTLAB_AVAILABLE:
            print("ReportLab not installed. Using HTML reports instead.")
        self.styles = None
        if REPORTLAB_AVAILABLE:
            self.styles = getSampleStyleSheet()
    
    def generate_daily_report(self, df_results: pd.DataFrame, 
                              summary: Dict,
                              parameters: Dict,
                              filename: str = None) -> Tuple[bytes, str]:
        """Generate a comprehensive daily report."""
        buffer = io.BytesIO()
        
        if REPORTLAB_AVAILABLE:
            self._generate_pdf_report(buffer, df_results, summary, parameters)
            content_type = "application/pdf"
        else:
            html_content = generate_html_report(df_results, summary, parameters)
            buffer.write(html_content.encode())
            content_type = "text/html"
        
        buffer.seek(0)
        
        if filename is None:
            filename = f"microgrid_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        return buffer.getvalue(), content_type
    
    def _generate_pdf_report(self, buffer: io.BytesIO, 
                           df_results: pd.DataFrame,
                           summary: Dict,
                           parameters: Dict):
        """Generate PDF report using ReportLab."""
        doc = SimpleDocTemplate(
            buffer, pagesize=letter,
            rightMargin=72, leftMargin=72,
            topMargin=72, bottomMargin=72
        )
        
        elements = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle', parent=self.styles['Heading1'],
            fontSize=24, spaceAfter=30,
            textColor=colors.HexColor("#1A365D")
        )
        elements.append(Paragraph("⚡ Optimal Grid Solutions", title_style))
        
        subtitle_style = ParagraphStyle(
            'Subtitle', parent=self.styles['Heading2'],
            fontSize=16, spaceAfter=20,
            textColor=colors.HexColor("#48BB78")
        )
        elements.append(Paragraph(
            f"Daily Optimization Report - {datetime.now().strftime('%B %d, %Y')}",
            subtitle_style
        ))
        elements.append(Spacer(1, 20))
        
        # Summary Table
        summary_data = [
            ["Total Cost", f"₹{summary.get('total_cost', 0):.2f}"],
            ["CO₂ Emissions", f"{summary.get('total_emissions', 0):.2f} kg"],
            ["Grid Usage", f"{summary.get('total_grid_usage', 0):.2f} kWh"],
            ["Renewable %", f"{summary.get('renewable_percentage', 0):.1f}%"]
        ]
        
        t = Table(summary_data, colWidths=[3*inch, 2*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#f8f9fa")),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor("#2c3e50")),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))
        
        # Footer
        elements.append(Spacer(1, 30))
        footer_style = ParagraphStyle('Footer', parent=self.styles['Normal'], fontSize=10, textColor=colors.grey)
        elements.append(Paragraph(
            f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | {COMPANY_BRANDING['footer_text']}",
            footer_style
        ))
        
        doc.build(elements)


def generate_quick_report(df_results: pd.DataFrame, summary: Dict, 
                         parameters: Dict) -> Tuple[bytes, str, str]:
    """Generate a quick report."""
    generator = PDFReportGenerator()
    data, content_type = generator.generate_daily_report(df_results, summary, parameters)
    filename = f"microgrid_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
    return data, content_type, filename


def generate_html_report(df_results: pd.DataFrame, summary: Dict,
                        parameters: Dict) -> str:
    """Generate enhanced HTML report with charts."""
    
    # Generate charts
    fig_power = go.Figure()
    fig_power.add_trace(go.Scatter(
        x=df_results["Hour"], y=df_results["Load (kW)"],
        mode='lines+markers', name='Load',
        line=dict(color='#2C3E50', width=2)
    ))
    fig_power.add_trace(go.Scatter(
        x=df_results["Hour"], y=df_results["Solar (kW)"],
        mode='lines', name='Solar', fill='tozeroy',
        fillcolor='rgba(244, 208, 63, 0.3)',
        line=dict(color='#F4D03F')
    ))
    fig_power.add_trace(go.Scatter(
        x=df_results["Hour"], y=df_results["Wind (kW)"],
        mode='lines', name='Wind', fill='tozeroy',
        fillcolor='rgba(52, 152, 219, 0.3)',
        line=dict(color='#3498DB')
    ))
    fig_power.update_layout(
        title="24-Hour Power Balance",
        xaxis_title="Hour",
        yaxis_title="Power (kW)",
        hovermode="x unified"
    )
    
    chart_html = fig_power.to_html(full_html=False, include_plotlyjs='cdn')
    
    recommendations = _generate_recommendations(summary, parameters)
    
    html = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Optimal Grid Solutions - Energy Report</title>
        <style>
            {get_branding_css()}
            body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 0; padding: 20px; background: #f5f5f5; }}
            .container {{ max-width: 1200px; margin: 0 auto; background: white; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
            .header {{ background: linear-gradient(135deg, #1A365D, #2D4A77); color: white; padding: 30px; border-radius: 10px 10px 0 0; text-align: center; }}
            .header-with-logo {{ background: linear-gradient(135deg, #1A365D, #2D4A77); padding: 20px; border-radius: 10px 10px 0 0; display: flex; align-items: center; justify-content: center; gap: 20px; }}
            .header-text {{ text-align: left; }}
            .header-text h1 {{ margin: 0; font-size: 28px; color: #ffffff; }}
            .header-text p {{ margin: 5px 0 0 0; color: #48BB78; font-size: 14px; }}
            .content {{ padding: 30px; }}
            .metrics {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 20px 0; }}
            .metric-card {{ background: linear-gradient(135deg, #f8f9fa, #fff); padding: 20px; border-radius: 10px; text-align: center; box-shadow: 0 2px 5px rgba(0,0,0,0.1); border-left: 4px solid #48BB78; }}
            .metric-value {{ font-size: 28px; font-weight: bold; color: #1A365D; }}
            .metric-label {{ color: #666; margin-top: 5px; }}
            table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
            th {{ background-color: #1A365D; color: white; }}
            tr:nth-child(even) {{ background-color: #f8f9fa; }}
            .recommendations {{ background: #e8f5e9; border-left: 4px solid #48BB78; padding: 15px; margin: 20px 0; border-radius: 0 10px 10px 0; }}
            .chart {{ margin: 30px 0; padding: 20px; background: #f8f9fa; border-radius: 10px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header-with-logo">
                <div style="flex-shrink: 0;">{SVG_LOGO_SMALL}</div>
                <div class="header-text">
                    <h1>{COMPANY_BRANDING['company_name']}</h1>
                    <p>{COMPANY_BRANDING['report_title']} | {datetime.now().strftime('%B %d, %Y')}</p>
                </div>
            </div>
            <div class="content">
                <h2>📊 Executive Summary</h2>
                <div class="metrics">
                    <div class="metric-card">
                        <div class="metric-value">₹{summary.get('total_cost', 0):.2f}</div>
                        <div class="metric-label">Total Cost</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value">{summary.get('total_emissions', 0):.1f} kg</div>
                        <div class="metric-label">CO₂ Emissions</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value">{summary.get('total_grid_usage', 0):.1f} kWh</div>
                        <div class="metric-label">Grid Usage</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value">{summary.get('renewable_percentage', 0):.1f}%</div>
                        <div class="metric-label">Renewable Share</div>
                    </div>
                </div>
                
                <h2>⚙️ System Parameters</h2>
                <table>
                    <tr><th>Parameter</th><th>Value</th></tr>
                    <tr><td>Battery Capacity</td><td>{parameters.get('battery_size', 100)} kWh</td></tr>
                    <tr><td>Base Price</td><td>₹{parameters.get('base_price', 5)}/unit</td></tr>
                    <tr><td>Peak Price</td><td>₹{parameters.get('peak_price', 15)}/unit</td></tr>
                    <tr><td>Grid Safety Limit</td><td>{parameters.get('grid_safety_limit', 50)} kW</td></tr>
                </table>
                
                <div class="chart">{chart_html}</div>
                
                <h2>📋 Hourly Performance</h2>
                <table>
                    <tr><th>Hour</th><th>Load</th><th>Solar</th><th>Wind</th><th>Grid</th><th>Cost</th></tr>
    """
    
    for _, row in df_results.iterrows():
        html += f"""
                    <tr>
                        <td>{int(row['Hour']):02d}:00</td>
                        <td>{row['Load (kW)']:.1f}</td>
                        <td>{row['Solar (kW)']:.1f}</td>
                        <td>{row['Wind (kW)']:.1f}</td>
                        <td>{row['Grid Usage (kW)']:.1f}</td>
                        <td>₹{row['Hourly Cost (₹)']:.2f}</td>
                    </tr>
        """
    
    html += """
                </table>
                
                <h2>💡 Recommendations</h2>
                <div class="recommendations">
                    <ul>
    """
    
    for rec in recommendations:
        html += f"<li>{rec}</li>"
    
    html += f"""
                    </ul>
                </div>
                {get_footer_html()}
            </div>
    </body>
    </html>
    """
    
    return html


def _generate_recommendations(summary: Dict, parameters: Dict) -> List[str]:
    """Generate recommendations based on analysis."""
    recommendations = []
    renewable_pct = summary.get('renewable_percentage', 0)
    total_cost = summary.get('total_cost', 0)
    total_emissions = summary.get('total_emissions', 0)
    grid_usage = summary.get('total_grid_usage', 0)
    
    if renewable_pct < 50:
        recommendations.append("Consider increasing renewable energy capacity to reduce carbon footprint.")
    elif renewable_pct < 70:
        recommendations.append("Good renewable penetration! Explore battery storage expansion for further optimization.")
    else:
        recommendations.append("Excellent renewable energy integration! Maintain current system configuration.")
    
    if total_cost > 2000:
        recommendations.append("High energy costs detected. Review peak hour usage and consider load shifting.")
    
    if total_emissions > 100:
        recommendations.append("Consider carbon offset programs or additional renewable capacity.")
    
    if grid_usage > 500:
        recommendations.append("High grid dependency. Evaluate increasing battery capacity for better autonomy.")
    
    return recommendations


def generate_text_report(df_results: pd.DataFrame, summary: Dict,
                         parameters: Dict) -> str:
    """Generate a simple text report."""
    report = f"""
================================================================================
                    SMART MICROGRID MANAGER PRO
                    Daily Optimization Report
================================================================================
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

                         EXECUTIVE SUMMARY
--------------------------------------------------------------------------------
Total Cost:           ₹{summary.get('total_cost', 0):.2f}
CO₂ Emissions:        {summary.get('total_emissions', 0):.2f} kg
Grid Usage:           {summary.get('total_grid_usage', 0):.2f} kWh
Renewable Percentage: {summary.get('renewable_percentage', 0):.1f}%

                         SYSTEM PARAMETERS
--------------------------------------------------------------------------------
Battery Capacity:     {parameters.get('battery_size', 100)} kWh
Base Price:           ₹{parameters.get('base_price', 5)}/unit
Peak Price:          ₹{parameters.get('peak_price', 15)}/unit
Grid Safety Limit:    {parameters.get('grid_safety_limit', 50)} kW

                         HOURLY DATA
--------------------------------------------------------------------------------
Hour    Load(kW)  Solar(kW)  Wind(kW)  Grid(kW)  Cost(₹)
--------------------------------------------------------------------------------
"""
    
    for _, row in df_results.iterrows():
        report += f"{int(row['Hour']):02d}:00   {row['Load (kW)']:7.1f}  {row['Solar (kW)']:8.1f}  {row['Wind (kW)']:8.1f}  {row['Grid Usage (kW)']:8.1f}  ₹{row['Hourly Cost (₹)']:7.2f}\n"
    
    recommendations = _generate_recommendations(summary, parameters)
    report += """
                         RECOMMENDATIONS
--------------------------------------------------------------------------------
"""
    for rec in recommendations:
        report += f"• {rec}\n"
    
    report += """
================================================================================
                         END OF REPORT
================================================================================
"""
    
    return report
